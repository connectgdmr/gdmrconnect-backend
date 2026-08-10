"""
routes/payroll.py — GDMR Connect
===================================
Payroll: salary structures, payslip generation, XLSX/PDF export, loans & advances.
"""
import io
import calendar
import threading
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, send_file
from bson import ObjectId

from database import (salary_structures_col, payslips_col, payroll_loans_col, users_col)
from decorators import token_required
from helpers import _is_admin, _to_money

bp = Blueprint("payroll", __name__)


# ── Component field groups ────────────────────────────────────────────────────

SALARY_EARNINGS   = ["basic", "da", "hra", "travel_allowance", "other_allowance"]
SALARY_DEDUCTIONS = ["pf", "esi", "lop", "tds", "other_deductions", "professional_tax", "gratuity"]
SALARY_DISPLAY_FIELDS = [
    "employee_code", "designation", "grade_profile",
    "days_worked", "bank_detail", "transaction_detail", "transaction_date",
]


def _payroll_allowed(user):
    if _is_admin(user):
        return True
    dept = (user.get("department") or "").strip().lower()
    return dept.startswith("accounts")


# ── Salary structures ─────────────────────────────────────────────────────────

@bp.route("/api/admin/payroll/salaries", methods=["GET"])
@token_required
def list_salaries():
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    employees  = list(users_col.find({"role": {"$in": ["employee", "manager", "owner"]}}, {"name": 1, "department": 1}))
    struct_map = {s["employee_id"]: s for s in salary_structures_col.find()}

    rows = []
    for e in employees:
        eid    = str(e["_id"])
        s      = struct_map.get(eid)
        salary = None
        if s:
            salary = {f: s.get(f, 0) for f in SALARY_EARNINGS + SALARY_DEDUCTIONS}
            salary["bonus"] = _to_money(s.get("bonus", 0))
            for df in SALARY_DISPLAY_FIELDS:
                salary[df] = s.get(df)
        rows.append({
            "employee_id":   eid,
            "employee_name": e.get("name", ""),
            "department":    e.get("department", ""),
            "salary":        salary,
        })
    return jsonify(rows), 200


@bp.route("/api/admin/payroll/salaries/<employee_id>", methods=["PUT"])
@token_required
def upsert_salary(employee_id):
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    try:
        emp = users_col.find_one({"_id": ObjectId(employee_id)})
    except Exception:
        return jsonify({"message": "Invalid employee ID"}), 400
    if not emp:
        return jsonify({"message": "Employee not found"}), 404

    data          = request.json or {}
    salary_fields = {f: _to_money(data.get(f, 0)) for f in SALARY_EARNINGS + SALARY_DEDUCTIONS}
    salary_fields["bonus"] = _to_money(data.get("bonus", 0))

    display_fields = {}
    for df in SALARY_DISPLAY_FIELDS:
        if df in data:
            display_fields[df] = data[df]
    display_fields.setdefault("transaction_detail", "Complete")

    now                 = datetime.now(timezone.utc)
    effective_date_raw  = data.get("effective_date")
    if effective_date_raw:
        try:
            effective_date = datetime.strptime(str(effective_date_raw)[:10], "%Y-%m-%d")
        except ValueError:
            return jsonify({"message": "Invalid effective_date. Use YYYY-MM-DD."}), 400
    else:
        effective_date = now

    history_entry = {
        **salary_fields,
        "effective_date":   effective_date,
        "increment_type":   str(data.get("increment_type",   "")).strip(),
        "increment_reason": str(data.get("increment_reason", "")).strip(),
        "recorded_at":      now,
    }

    salary_structures_col.update_one(
        {"employee_id": employee_id},
        {
            "$set":  {**salary_fields, **display_fields, "employee_id": employee_id, "updated_at": now},
            "$push": {"salary_history": history_entry},
        },
        upsert=True,
    )
    return jsonify({"message": "Salary structure saved", "salary": {**salary_fields, **display_fields}}), 200


@bp.route("/api/admin/payroll/salaries/<employee_id>/history", methods=["GET"])
@token_required
def get_salary_history(employee_id):
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    try:
        emp = users_col.find_one({"_id": ObjectId(employee_id)}, {"name": 1})
    except Exception:
        return jsonify({"message": "Invalid employee ID"}), 400
    if not emp:
        return jsonify({"message": "Employee not found"}), 404

    struct = salary_structures_col.find_one({"employee_id": employee_id})
    if not struct:
        return jsonify([]), 200

    history = struct.get("salary_history", [])
    history.sort(key=lambda h: h.get("effective_date") or datetime.min)

    result = []
    for entry in history:
        row = {f: entry.get(f, 0) for f in SALARY_EARNINGS + SALARY_DEDUCTIONS}
        row["bonus"] = _to_money(entry.get("bonus", 0))
        ed = entry.get("effective_date")
        row["effective_date"]   = ed.isoformat() if isinstance(ed, datetime) else ed
        row["increment_type"]   = entry.get("increment_type", "")
        row["increment_reason"] = entry.get("increment_reason", "")
        rec = entry.get("recorded_at")
        row["recorded_at"]      = rec.isoformat() if isinstance(rec, datetime) else rec
        result.append(row)

    return jsonify(result), 200


# ── Payroll run ───────────────────────────────────────────────────────────────

@bp.route("/api/admin/payroll/run", methods=["POST"])
@token_required
def run_payroll():
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    data = request.json or {}
    try:
        month = int(data.get("month"))
        year  = int(data.get("year"))
    except (TypeError, ValueError):
        return jsonify({"message": "month (1-12) and year are required"}), 400
    if not (1 <= month <= 12) or year < 2000:
        return jsonify({"message": "Invalid month or year"}), 400

    period = datetime(year, month, 1).strftime("%B %Y")
    now    = datetime.now(timezone.utc)

    adj_map: dict = {}
    for adj in (data.get("adjustments") or []):
        eid_adj = str(adj.get("employee_id") or "").strip()
        if eid_adj:
            adj_map[eid_adj] = adj

    created = 0
    skipped = 0
    for s in salary_structures_col.find():
        eid = s.get("employee_id")
        if not eid:
            continue

        if payslips_col.find_one({"employee_id": eid, "month": month, "year": year}):
            skipped += 1
            continue

        try:
            emp = users_col.find_one({"_id": ObjectId(eid)})
        except Exception:
            emp = None
        if not emp:
            continue

        earnings = {f: _to_money(s.get(f, 0)) for f in SALARY_EARNINGS}
        bonus    = _to_money(s.get("bonus", 0))
        gross    = round(sum(earnings.values()), 2)

        adj               = adj_map.get(eid, {})
        pf                = _to_money(s.get("pf", 0))
        professional_tax  = _to_money(s.get("professional_tax", 0))
        gratuity          = _to_money(s.get("gratuity", 0))
        esi               = _to_money(adj.get("esi", 0))
        lop               = _to_money(adj.get("lop", 0))
        tds               = _to_money(adj.get("tds", 0))
        other_deductions  = _to_money(adj.get("other_deductions", 0))
        remark            = str(adj.get("remark", "") or "").strip()
        total_deductions  = round(pf + esi + lop + tds + other_deductions + professional_tax + gratuity, 2)
        net               = round(gross + bonus - total_deductions, 2)

        display = {df: s.get(df) for df in SALARY_DISPLAY_FIELDS}
        if not display.get("transaction_detail"):
            display["transaction_detail"] = "Complete"
        if not display.get("designation"):
            display["designation"] = emp.get("position", "")
        display["employee_code"] = emp.get("employee_code") or display.get("employee_code") or ""

        loan_emi         = 0.0
        advance_recovery = 0.0
        for loan in list(payroll_loans_col.find({"employee_id": eid, "status": "active"})):
            lid         = loan["_id"]
            outstanding = _to_money(loan.get("outstanding", 0))
            emi         = _to_money(loan.get("emi_per_month", 0))
            loan_type   = loan.get("type", "loan")
            if outstanding <= 0:
                payroll_loans_col.update_one({"_id": lid}, {"$set": {"status": "closed", "outstanding": 0}})
                continue
            if loan_type == "advance":
                deduct = min(outstanding, emi or outstanding)
                advance_recovery = round(advance_recovery + deduct, 2)
            else:
                deduct   = min(outstanding, emi)
                loan_emi = round(loan_emi + deduct, 2)
            new_outstanding = round(outstanding - deduct, 2)
            loan_update = {"outstanding": new_outstanding}
            if new_outstanding <= 0:
                loan_update["status"] = "closed"
            payroll_loans_col.update_one({"_id": lid}, {"$set": loan_update})

        net = round(gross + bonus - total_deductions - loan_emi - advance_recovery, 2)

        payslip = {
            "employee_id":      eid,
            "employee_name":    emp.get("name", ""),
            "department":       emp.get("department", ""),
            "month":            month,
            "year":             year,
            "period":           period,
            **earnings,
            "bonus":            bonus,
            "gross":            gross,
            "pf":               pf,
            "esi":              esi,
            "lop":              lop,
            "tds":              tds,
            "other_deductions": other_deductions,
            "professional_tax": professional_tax,
            "gratuity":         gratuity,
            "total_deductions": total_deductions,
            "loan_emi":         loan_emi,
            "advance_recovery": advance_recovery,
            "net":              net,
            "remark":           remark,
            **display,
            "status":           "Pending",
            "created_at":       now,
        }
        try:
            payslips_col.insert_one(payslip)
            created += 1
        except Exception:
            skipped += 1

    return jsonify({
        "message": f"Payroll run complete for {period}.",
        "period":  period,
        "created": created,
        "skipped": skipped,
    }), 200


# ── Payslip listing ───────────────────────────────────────────────────────────

@bp.route("/api/admin/payroll/payslips", methods=["GET"])
@token_required
def list_payslips():
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    query = {}
    month = request.args.get("month")
    year  = request.args.get("year")
    # Full history for one employee (used by the clickable-name payroll
    # profile popup) — deliberately not paired with month/year so it
    # returns every payslip they've ever had, newest first.
    employee_id = request.args.get("employee_id")
    if employee_id:
        query["employee_id"] = employee_id
    if month:
        try:
            query["month"] = int(month)
        except ValueError:
            return jsonify({"message": "Invalid month"}), 400
    if year:
        try:
            query["year"] = int(year)
        except ValueError:
            return jsonify({"message": "Invalid year"}), 400

    rows = []
    for p in payslips_col.find(query).sort([("year", -1), ("month", -1), ("employee_name", 1)]):
        p["_id"] = str(p["_id"])
        rows.append(p)
    return jsonify(rows), 200


# ── Export helpers ────────────────────────────────────────────────────────────

PAYROLL_EXPORT_COLUMNS = [
    ("SL No", "dark"), ("Employee ID", "dark"), ("DOJ", "dark"), ("Employee Name", "dark"),
    ("Designation", "dark"), ("Grade", "dark"), ("Department", "dark"),
    ("Current Salary", "salmon"), ("Annual Salary", "salmon"), ("Per day", "salmon"),
    ("LOP", "cyan"),
    ("Total Working Days", "salmon"),
    ("Salary as per days worked", "yellow"),
    ("Basic", "salmon"), ("DA", "salmon"), ("HRA", "salmon"), ("Travel Allowance", "salmon"),
    ("Other Allowance", "salmon"), ("Bonus", "salmon"),
    ("Gross Salary", "yellow"),
    ("LOP", "salmon"), ("PF", "salmon"), ("Esi", "salmon"), ("Professional Tax", "salmon"),
    ("Gratuity", "salmon"), ("TDS", "salmon"), ("Loan/Deduction", "salmon"),
    ("TOTAL DEDUCTIONS", "yellow"),
    ("Net Salary", "green"), ("Bank", "green"), ("Payment Status", "green"),
]
PAYROLL_EXPORT_FILL = {
    "dark":   "1C2B39",
    "salmon": "F4CCCC",
    "cyan":   "9FE2EA",
    "yellow": "FFF2CC",
    "green":  "D9EAD3",
}
PAYROLL_EXPORT_WIDE_COLS   = {"Employee Name", "Designation", "Department", "Bank"}
PAYROLL_EXPORT_NARROW_COLS = {"SL No", "Grade", "LOP"}


def _color_group_ranges(columns):
    ranges = []
    start  = 0
    for i in range(1, len(columns) + 1):
        if i == len(columns) or columns[i][1] != columns[start][1]:
            ranges.append((start, i - 1, columns[start][1]))
            start = i
    return ranges


def _build_payroll_export_rows(slips, days_in_month, doj_map):
    rows = []
    for i, p in enumerate(slips, start=1):
        current_salary = round(
            _to_money(p.get("basic")) + _to_money(p.get("da")) + _to_money(p.get("hra"))
            + _to_money(p.get("travel_allowance")) + _to_money(p.get("other_allowance")), 2
        )
        annual_salary             = round(current_salary * 12, 2)
        per_day                   = round(current_salary / days_in_month, 2) if days_in_month else 0
        lop_amount                = _to_money(p.get("lop"))
        lop_days                  = round(lop_amount / per_day, 2) if per_day else 0
        total_working_days        = round(days_in_month - lop_days, 2)
        salary_as_per_days_worked = round(current_salary - lop_amount, 2)
        bonus                     = _to_money(p.get("bonus"))
        gross_salary              = round(current_salary + bonus, 2)
        loan_deduction            = round(
            _to_money(p.get("other_deductions")) + _to_money(p.get("loan_emi")) + _to_money(p.get("advance_recovery")), 2
        )
        total_deductions = round(
            lop_amount + _to_money(p.get("pf")) + _to_money(p.get("esi")) + _to_money(p.get("professional_tax"))
            + _to_money(p.get("gratuity")) + _to_money(p.get("tds")) + loan_deduction, 2
        )
        net_salary = round(gross_salary - total_deductions, 2)

        rows.append([
            i,
            p.get("employee_code") or p.get("employee_id", ""),
            doj_map.get(p.get("employee_id"), ""),
            p.get("employee_name", ""),
            p.get("designation", ""),
            p.get("grade_profile", ""),
            p.get("department", ""),
            current_salary, annual_salary, per_day,
            lop_days,
            total_working_days,
            salary_as_per_days_worked,
            _to_money(p.get("basic")), _to_money(p.get("da")), _to_money(p.get("hra")),
            _to_money(p.get("travel_allowance")), _to_money(p.get("other_allowance")), bonus,
            gross_salary,
            lop_amount, _to_money(p.get("pf")), _to_money(p.get("esi")), _to_money(p.get("professional_tax")),
            _to_money(p.get("gratuity")), _to_money(p.get("tds")), loan_deduction,
            total_deductions,
            net_salary, p.get("bank_detail", ""), p.get("status", "Pending"),
        ])
    return rows


def _render_payroll_xlsx(month, year, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"[:31]

    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, (label, color_key) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=label)
        cell.fill       = PatternFill("solid", fgColor=PAYROLL_EXPORT_FILL[color_key])
        cell.font       = Font(bold=True, size=9, color="FFFFFF" if color_key == "dark" else "1C2B39")
        cell.alignment  = center_wrap
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=r, column=col_idx, value=val).font = Font(size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _render_payroll_pdf(month, year, rows):
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    def fmt_cell(v):
        if isinstance(v, float):
            return f"{v:,.2f}"
        return "" if v is None else str(v)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=8 * mm, rightMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
    )

    weights = [
        1.5 if label in PAYROLL_EXPORT_WIDE_COLS else (0.7 if label in PAYROLL_EXPORT_NARROW_COLS else 1.0)
        for label, _ in PAYROLL_EXPORT_COLUMNS
    ]
    total_weight = sum(weights)
    col_widths   = [doc.width * (w / total_weight) for w in weights]

    cell_style = ParagraphStyle("cell", fontSize=6, leading=7, alignment=1)

    def header_row():
        out = []
        for label, color_key in PAYROLL_EXPORT_COLUMNS:
            style = ParagraphStyle(
                "hdr", fontSize=6.5, leading=7.5, alignment=1, fontName="Helvetica-Bold",
                textColor=rl_colors.white if color_key == "dark" else rl_colors.HexColor("#1C2B39"),
            )
            out.append(Paragraph(label, style))
        return out

    body_rows = [[Paragraph(fmt_cell(v), cell_style) for v in row] for row in rows]
    title     = Paragraph(
        f"<b>Payroll Export — {calendar.month_name[month]} {year}</b>",
        getSampleStyleSheet()["Heading3"]
    )

    table = Table([header_row()] + body_rows, colWidths=col_widths, repeatRows=1)
    style = [
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",           (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#94A3B8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
        ("TOPPADDING",    (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2), ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
    ]
    for start, end, color_key in _color_group_ranges(PAYROLL_EXPORT_COLUMNS):
        style.append(("BACKGROUND", (start, 0), (end, 0), rl_colors.HexColor("#" + PAYROLL_EXPORT_FILL[color_key])))
    table.setStyle(TableStyle(style))

    doc.build([title, table])
    buf.seek(0)
    return buf


@bp.route("/api/admin/payroll/export", methods=["GET"])
@token_required
def export_payroll():
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    try:
        month = int(request.args.get("month"))
        year  = int(request.args.get("year"))
    except (TypeError, ValueError):
        return jsonify({"message": "month (1-12) and year are required"}), 400
    if not (1 <= month <= 12) or year < 2000:
        return jsonify({"message": "Invalid month or year"}), 400

    fmt = (request.args.get("format") or "xlsx").lower()
    if fmt not in ("xlsx", "pdf"):
        return jsonify({"message": "format must be 'xlsx' or 'pdf'"}), 400

    slips         = list(payslips_col.find({"month": month, "year": year}).sort([("department", 1), ("employee_name", 1)]))
    days_in_month = calendar.monthrange(year, month)[1]

    doj_map        = {}
    emp_object_ids = []
    for p in slips:
        try:
            emp_object_ids.append(ObjectId(p.get("employee_id")))
        except Exception:
            pass
    if emp_object_ids:
        for u in users_col.find({"_id": {"$in": emp_object_ids}}, {"doj": 1}):
            doj_map[str(u["_id"])] = u.get("doj") or ""

    rows         = _build_payroll_export_rows(slips, days_in_month, doj_map)
    period_label = f"{calendar.month_name[month]}_{year}"

    if fmt == "pdf":
        buf = _render_payroll_pdf(month, year, rows)
        return send_file(buf, mimetype="application/pdf", as_attachment=True,
                         download_name=f"Payroll_Export_{period_label}.pdf")

    buf = _render_payroll_xlsx(month, year, rows)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Payroll_Export_{period_label}.xlsx",
    )


@bp.route("/api/admin/payroll/payslips/<payslip_id>/status", methods=["PUT"])
@token_required
def update_payslip_status(payslip_id):
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    try:
        obj = ObjectId(payslip_id)
    except Exception:
        return jsonify({"message": "Invalid ID"}), 400

    data   = request.json or {}
    status = data.get("status")
    if status not in ("Pending", "Paid"):
        return jsonify({"message": "status must be 'Pending' or 'Paid'"}), 400

    update = {"status": status, "updated_at": datetime.now(timezone.utc)}
    if status == "Paid":
        update["paid_at"] = datetime.now(timezone.utc)

    result = payslips_col.update_one({"_id": obj}, {"$set": update})
    if result.matched_count == 0:
        return jsonify({"message": "Payslip not found"}), 404
    return jsonify({"message": f"Payslip marked {status}"}), 200


@bp.route("/api/my/payslips", methods=["GET"])
@token_required
def my_payslips():
    uid  = str(request.user["_id"])
    rows = []
    for p in payslips_col.find({"employee_id": uid}).sort([("year", -1), ("month", -1)]):
        p["_id"] = str(p["_id"])
        rows.append(p)
    return jsonify(rows), 200


# ── Loans & Advances ──────────────────────────────────────────────────────────

@bp.route("/api/admin/payroll/loans", methods=["GET"])
@token_required
def list_payroll_loans():
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    status_filter = request.args.get("status")
    query = {}
    if status_filter in ("active", "closed"):
        query["status"] = status_filter

    rows = []
    for loan in payroll_loans_col.find(query).sort([("created_at", -1)]):
        loan["_id"] = str(loan["_id"])
        rows.append(loan)
    return jsonify(rows), 200


@bp.route("/api/admin/payroll/loans", methods=["POST"])
@token_required
def create_payroll_loan():
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    data        = request.get_json(silent=True) or {}
    employee_id = (data.get("employee_id") or "").strip()
    loan_type   = (data.get("type") or "loan").strip().lower()
    if not employee_id:
        return jsonify({"message": "employee_id is required"}), 400
    if loan_type not in ("loan", "advance"):
        return jsonify({"message": "type must be 'loan' or 'advance'"}), 400

    try:
        emp = users_col.find_one({"_id": ObjectId(employee_id)})
    except Exception:
        emp = None
    if not emp:
        return jsonify({"message": "Employee not found"}), 404

    amount = _to_money(data.get("amount", 0))
    if amount <= 0:
        return jsonify({"message": "amount must be greater than 0"}), 400

    emi              = _to_money(data.get("emi_per_month", 0))
    repayment_months = data.get("repayment_months")
    try:
        repayment_months = int(repayment_months) if repayment_months else None
    except (ValueError, TypeError):
        repayment_months = None

    loan_doc = {
        "employee_id":      employee_id,
        "employee_name":    emp.get("name", ""),
        "department":       emp.get("department", ""),
        "type":             loan_type,
        "amount":           amount,
        "emi_per_month":    emi,
        "outstanding":      amount,
        "repayment_months": repayment_months,
        "disbursed_date":   data.get("disbursed_date", ""),
        "reason":           data.get("reason", ""),
        "notes":            data.get("notes", ""),
        "status":           "active",
        "created_at":       datetime.now(timezone.utc),
    }
    result           = payroll_loans_col.insert_one(loan_doc)
    loan_doc["_id"]  = str(result.inserted_id)
    return jsonify(loan_doc), 201


@bp.route("/api/admin/payroll/loans/<loan_id>/status", methods=["PUT"])
@token_required
def update_payroll_loan_status(loan_id):
    if not _payroll_allowed(request.user):
        return jsonify({"message": "Unauthorized"}), 403

    try:
        obj = ObjectId(loan_id)
    except Exception:
        return jsonify({"message": "Invalid ID"}), 400

    data   = request.get_json(silent=True) or {}
    status = (data.get("status") or "").strip().lower()
    if status not in ("active", "closed"):
        return jsonify({"message": "status must be 'active' or 'closed'"}), 400

    update = {"status": status}
    if status == "closed":
        update["outstanding"] = 0

    result = payroll_loans_col.update_one({"_id": obj}, {"$set": update})
    if result.matched_count == 0:
        return jsonify({"message": "Loan not found"}), 404
    return jsonify({"message": f"Loan marked {status}"}), 200


@bp.route("/api/my/loans", methods=["GET"])
@token_required
def my_loans():
    uid  = str(request.user["_id"])
    rows = []
    for loan in payroll_loans_col.find({"employee_id": uid}).sort([("created_at", -1)]):
        loan["_id"] = str(loan["_id"])
        rows.append(loan)
    return jsonify(rows), 200
