"""
routes/attendance_reports.py — GDMR Connect
==============================================
Two PDF report layouts modeled directly on the HR team's existing manual
spreadsheets, so switching to the app doesn't mean losing a familiar format:

  - Monthly Report: one row per employee, one column per day of a chosen
    month, each cell showing that day's status (Present / PL / HD / LOP /
    Not Checked In), a "Total leaves" column, and an employee Status
    column (Confirmed / Probation / Contract / Intern).
  - Master Tracker: one row per employee, one group of four columns
    (Total Working Days / Attended Days / Leaves / Percentage) per month
    of a chosen year, plus a yearly Total Percentage.

Both are built entirely from data already in this app (attendance_col,
leaves_col, holidays_col, users_col) — no new fields were added. Two
things the reference spreadsheets track that this app has no concept of
at all are intentionally NOT reproduced, rather than faked:
  - Sick Leave vs Planned Leave as distinct categories — every approved
    leave here is just "PL", there's no SL split.
  - Work-From-Home (WFH) — not a request type in this app, so it never
    appears; those days fall back to whatever they'd otherwise show
    (Present if checked in, LOP if not).
A leave counts as "approved" for these reports as soon as the employee's
MANAGER has approved it — it doesn't wait for Admin's final sign-off too,
per how HR actually treats it day to day (see _leave_is_approved below).
"""
import calendar
import csv
import io
from datetime import datetime, timezone

from flask import Blueprint, request, send_file

from database import attendance_col, leaves_col, users_col, holidays_col
from decorators import token_required
from helpers import (
    _is_admin, _has_module_grant, is_offboarded, is_weekend_day,
    get_company_holiday_dates, _date_str,
)
from config import IST

bp = Blueprint("attendance_reports", __name__)


def _reports_allowed(user):
    return _is_admin(user) or _has_module_grant(user, "attendance") or _has_module_grant(user, "summary")


def _leave_is_approved(leave):
    """A leave counts as approved for reporting the moment the employee's
    manager has signed off — it doesn't wait for Admin's separate final
    approval too. (Admin-submitted leaves have no manager in the loop at
    all, so those go by the overall status instead.)"""
    return leave.get("manager_status") == "Approved" or leave.get("status") == "Approved"


def _report_employee_status(emp):
    """Confirmed / Probation / Contract / Intern — derived, not stored
    anywhere: Contract/Internship employment_type map straight across,
    everyone else is Probation until their confirmation_date has passed."""
    et = emp.get("employment_type") or "Permanent"
    if et == "Contract":
        return "Contract"
    if et == "Internship":
        return "Intern"
    conf = _date_str(emp.get("confirmation_date"))
    today = str(datetime.now(IST).date())
    return "Confirmed" if conf and conf <= today else "Probation"


def _active_employees():
    rows = list(users_col.find(
        {"role": {"$in": ["employee", "manager"]}},
        {"name": 1, "employee_code": 1, "department": 1, "employment_type": 1,
         "confirmation_date": 1, "doj": 1, "resignation": 1},
    ))
    return [e for e in rows if not is_offboarded(e)]


def _dept_of(emp):
    d = emp.get("department")
    return ", ".join(d) if isinstance(d, list) else (d or "—")


# ── Monthly Report ──────────────────────────────────────────────────────────

def _build_monthly_report_rows(month, year):
    days_in_month = calendar.monthrange(year, month)[1]
    month_str     = f"{year:04d}-{month:02d}"
    today_str     = str(datetime.now(IST).date())
    holiday_dates = get_company_holiday_dates()
    holiday_names = {h["date"]: h.get("name") for h in holidays_col.find(
        {"date": {"$in": list(holiday_dates)}}, {"date": 1, "name": 1}
    )}

    employees = _active_employees()
    uids      = [str(e["_id"]) for e in employees]

    checkins = {}  # {uid: {date_str}}
    for rec in attendance_col.find(
        {"user_id": {"$in": uids}, "type": "checkin", "date": {"$regex": f"^{month_str}"}},
        {"user_id": 1, "date": 1},
    ):
        checkins.setdefault(rec["user_id"], set()).add(rec["date"])

    month_start = f"{month_str}-01"
    month_end   = f"{month_str}-{days_in_month:02d}"
    leaves_by_uid = {}
    for lv in leaves_col.find({
        "user_id": {"$in": uids},
        "from_date": {"$lte": month_end}, "to_date": {"$gte": month_start},
        "status": {"$nin": ["Rejected", "Cancelled"]},
    }):
        if _leave_is_approved(lv):
            leaves_by_uid.setdefault(lv["user_id"], []).append(lv)

    day_headers, dow_headers, day_is_off = [], [], []
    for d in range(1, days_in_month + 1):
        day_str = f"{month_str}-{d:02d}"
        dt = datetime(year, month, d)
        day_headers.append(str(d))
        dow_headers.append(dt.strftime("%a"))
        day_is_off.append(is_weekend_day(day_str, holiday_dates))

    rows = []
    for emp in sorted(employees, key=lambda e: (e.get("name") or "")):
        uid = str(emp["_id"])
        joined = _date_str(emp.get("doj"))
        lwd    = _date_str((emp.get("resignation") or {}).get("last_working_day"))
        emp_checkins = checkins.get(uid, set())
        emp_leaves   = leaves_by_uid.get(uid, [])

        cells = []
        total_leave_days = 0.0
        for d in range(1, days_in_month + 1):
            day_str = f"{month_str}-{d:02d}"
            if day_is_off[d - 1]:
                cells.append(holiday_names.get(day_str, ""))
                continue
            if (joined and day_str < joined) or (lwd and day_str > lwd):
                cells.append("")
                continue
            if day_str in emp_checkins:
                cells.append("Present")
                continue
            leave_today = next((lv for lv in emp_leaves if lv.get("from_date", "") <= day_str <= lv.get("to_date", "")), None)
            if leave_today:
                is_half = leave_today.get("type") == "half"
                total_leave_days += 0.5 if is_half else 1.0
                cells.append("HD" if is_half else "PL")
                continue
            if day_str == today_str:
                cells.append("Not Checked In")
            elif day_str > today_str:
                cells.append("")
            else:
                cells.append("LOP")

        rows.append({
            "employee_code": emp.get("employee_code") or "—",
            "name":          emp.get("name") or "",
            "department":    _dept_of(emp),
            "status":        _report_employee_status(emp),
            "total_leaves":  total_leave_days,
            "cells":         cells,
        })

    return rows, day_headers, dow_headers, day_is_off


def _render_monthly_report_pdf(rows, day_headers, dow_headers, day_is_off, month, year):
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    n_days = len(day_headers)
    # Custom wide page — 31 day columns plus 5 fixed columns needs far more
    # width than any standard sheet size gives; height stays A3-landscape-ish.
    page_w = (34 * 5 + 13 * n_days + 20) * mm
    page_h = 297 * mm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=(page_w, page_h),
        leftMargin=6 * mm, rightMargin=6 * mm, topMargin=8 * mm, bottomMargin=8 * mm,
    )

    fixed_labels = ["Employee ID", "Employee Name", "Department", "Status", "Total Leaves"]
    fixed_weights = [1.1, 1.8, 1.5, 1.0, 0.9]
    day_weight = 0.55
    weights = fixed_weights + [day_weight] * n_days
    total_weight = sum(weights)
    col_widths = [doc.width * (w / total_weight) for w in weights]

    hdr_style   = ParagraphStyle("hdr", fontSize=6.5, leading=7.5, alignment=1, fontName="Helvetica-Bold")
    cell_style  = ParagraphStyle("cell", fontSize=5.5, leading=6.5, alignment=1)

    row1 = [""] * 5 + [Paragraph(d, hdr_style) for d in day_headers]
    row2 = [Paragraph(l, hdr_style) for l in fixed_labels] + [Paragraph(d, hdr_style) for d in dow_headers]

    STATUS_COLOR = {
        "Present": rl_colors.HexColor("#16a34a"),
        "PL":      rl_colors.HexColor("#d97706"),
        "HD":      rl_colors.HexColor("#d97706"),
        "LOP":     rl_colors.HexColor("#dc2626"),
        "Not Checked In": rl_colors.HexColor("#94a3b8"),
    }
    body_rows = []
    for r in rows:
        row = [
            Paragraph(r["employee_code"], cell_style),
            Paragraph(r["name"], cell_style),
            Paragraph(r["department"], cell_style),
            Paragraph(r["status"], cell_style),
            Paragraph(f'{r["total_leaves"]:g}', cell_style),
        ]
        for c in r["cells"]:
            color = STATUS_COLOR.get(c)
            style = cell_style if not color else ParagraphStyle("cellc", parent=cell_style, textColor=color, fontName="Helvetica-Bold")
            row.append(Paragraph(c, style))
        body_rows.append(row)

    title = Paragraph(f"<b>Monthly Attendance Report — {calendar.month_name[month]} {year}</b>", getSampleStyleSheet()["Heading3"])
    table = Table([row1, row2] + body_rows, colWidths=col_widths, repeatRows=2)

    style = [
        ("ALIGN",  (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",   (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 1), rl_colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0, 0), (-1, 1), rl_colors.white),
        ("ROWBACKGROUNDS", (0, 2), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.5), ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
        ("SPAN", (0, 0), (0, 1)), ("SPAN", (1, 0), (1, 1)), ("SPAN", (2, 0), (2, 1)),
        ("SPAN", (3, 0), (3, 1)), ("SPAN", (4, 0), (4, 1)),
    ]
    # Weekend/holiday day columns get the same soft-yellow highlight the
    # reference sheet uses, across every row including the header.
    for i, off in enumerate(day_is_off):
        if off:
            col = 5 + i
            style.append(("BACKGROUND", (col, 2), (col, -1), rl_colors.HexColor("#fef9c3")))
    table.setStyle(TableStyle(style))

    doc.build([title, table])
    buf.seek(0)
    return buf


@bp.route("/api/admin/reports/monthly-attendance-pdf", methods=["GET"])
@token_required
def monthly_attendance_pdf():
    if not _reports_allowed(request.user):
        return "Unauthorized", 403
    try:
        month = int(request.args.get("month"))
        year  = int(request.args.get("year"))
    except (TypeError, ValueError):
        return "month (1-12) and year are required", 400
    if not (1 <= month <= 12) or year < 2000:
        return "Invalid month or year", 400

    rows, day_headers, dow_headers, day_is_off = _build_monthly_report_rows(month, year)
    buf = _render_monthly_report_pdf(rows, day_headers, dow_headers, day_is_off, month, year)
    filename = f"Monthly_Attendance_Report_{calendar.month_name[month]}_{year}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


def _render_monthly_report_csv(rows, day_headers):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Employee ID", "Employee Name", "Department", "Status", "Total Leaves"] + day_headers)
    for r in rows:
        w.writerow([r["employee_code"], r["name"], r["department"], r["status"], f'{r["total_leaves"]:g}'] + r["cells"])
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))  # BOM so Excel opens it without mangling names
    out.seek(0)
    return out


@bp.route("/api/admin/reports/monthly-attendance-csv", methods=["GET"])
@token_required
def monthly_attendance_csv():
    if not _reports_allowed(request.user):
        return "Unauthorized", 403
    try:
        month = int(request.args.get("month"))
        year  = int(request.args.get("year"))
    except (TypeError, ValueError):
        return "month (1-12) and year are required", 400
    if not (1 <= month <= 12) or year < 2000:
        return "Invalid month or year", 400

    rows, day_headers, _dow, _off = _build_monthly_report_rows(month, year)
    buf = _render_monthly_report_csv(rows, day_headers)
    filename = f"Monthly_Attendance_Report_{calendar.month_name[month]}_{year}.csv"
    return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


# ── Master Tracker (yearly) ─────────────────────────────────────────────────

def _month_working_stats(uid, year, month, holiday_dates, leaves_for_emp, joined, lwd):
    """(total_working_days, attended_days, leave_days) for one employee/month."""
    days_in_month = calendar.monthrange(year, month)[1]
    today_str     = str(datetime.now(IST).date())
    month_str     = f"{year:04d}-{month:02d}"

    checkin_dates = {
        rec["date"] for rec in attendance_col.find(
            {"user_id": uid, "type": "checkin", "date": {"$regex": f"^{month_str}"}}, {"date": 1}
        )
    }

    total, attended, leave_days = 0, 0.0, 0.0
    for d in range(1, days_in_month + 1):
        day_str = f"{month_str}-{d:02d}"
        if day_str > today_str:
            continue
        if is_weekend_day(day_str, holiday_dates):
            continue
        if (joined and day_str < joined) or (lwd and day_str > lwd):
            continue
        total += 1
        if day_str in checkin_dates:
            attended += 1
            continue
        leave_today = next((lv for lv in leaves_for_emp if lv.get("from_date", "") <= day_str <= lv.get("to_date", "")), None)
        if leave_today:
            is_half = leave_today.get("type") == "half"
            leave_days += 0.5 if is_half else 1.0
            if is_half:
                attended += 0.5
        # else: LOP — counts toward total (worked-day expected) but not attended

    return total, attended, leave_days


def _build_master_tracker_rows(year):
    holiday_dates = get_company_holiday_dates()
    employees = _active_employees()
    uids      = [str(e["_id"]) for e in employees]

    year_start = f"{year:04d}-01-01"
    year_end   = f"{year:04d}-12-31"
    leaves_by_uid = {}
    for lv in leaves_col.find({
        "user_id": {"$in": uids},
        "from_date": {"$lte": year_end}, "to_date": {"$gte": year_start},
        "status": {"$nin": ["Rejected", "Cancelled"]},
    }):
        if _leave_is_approved(lv):
            leaves_by_uid.setdefault(lv["user_id"], []).append(lv)

    rows = []
    for emp in sorted(employees, key=lambda e: (e.get("name") or "")):
        uid    = str(emp["_id"])
        joined = _date_str(emp.get("doj"))
        lwd    = _date_str((emp.get("resignation") or {}).get("last_working_day"))
        emp_leaves = leaves_by_uid.get(uid, [])

        months = []
        yr_total, yr_attended = 0.0, 0.0
        for m in range(1, 13):
            total, attended, leave_days = _month_working_stats(uid, year, m, holiday_dates, emp_leaves, joined, lwd)
            pct = round(attended / total * 100, 1) if total else None
            months.append({"total": total, "attended": attended, "leaves": leave_days, "pct": pct})
            yr_total += total
            yr_attended += attended

        rows.append({
            "employee_code": emp.get("employee_code") or "—",
            "name":          emp.get("name") or "",
            "department":    _dept_of(emp),
            "months":        months,
            "year_pct":      round(yr_attended / yr_total * 100, 1) if yr_total else None,
        })
    return rows


# The Master Tracker download is a four-part report that mirrors the tabs of
# the HR team's manual spreadsheet:
#   1. Master Sheet     — Total / Attended / Leaves / % per month (the original)
#   2. Dep wise         — monthly attendance % per employee, grouped by
#                         department, with per-department Average % / Quarter %
#   3. Leave Monitoring — per month, per employee: Monday-leave / Friday-leave /
#                         HD / Early-leaving / WFH / Late-coming counts
#   4. Late Coming      — the actual dates each employee checked in late, by
#                         month, plus a yearly count
# The PDF puts each on its own (very wide) page; the Excel export puts each on
# its own worksheet. Neither adds any data model — everything comes from
# attendance_col.status_indicator / day_type and approved leaves.


def _dept_wise_blocks_from_tracker(tracker_rows):
    """Regroup the already-computed Master-Sheet rows by department. Each block:
    {department, employees:[{employee_code, name, pcts:[12]}], average:[12], quarter:[12]}
    where `quarter` carries the Q1..Q4 mean of the department's monthly averages
    in the Jan/Apr/Jul/Oct slots (blank elsewhere), matching the reference sheet."""
    groups = {}
    for r in tracker_rows:
        groups.setdefault(r["department"], []).append(r)

    blocks = []
    for dept in sorted(groups):
        emp_rows = [
            {"employee_code": r["employee_code"], "name": r["name"],
             "pcts": [mo["pct"] for mo in r["months"]]}
            for r in sorted(groups[dept], key=lambda r: r["name"].lower())
        ]
        average = []
        for mi in range(12):
            vals = [er["pcts"][mi] for er in emp_rows if er["pcts"][mi] is not None]
            average.append(round(sum(vals) / len(vals), 1) if vals else None)
        quarter = [None] * 12
        for qi in range(4):
            qv = [average[qi * 3 + k] for k in range(3) if average[qi * 3 + k] is not None]
            quarter[qi * 3] = round(sum(qv) / len(qv), 1) if qv else None
        blocks.append({"department": dept, "employees": emp_rows,
                       "average": average, "quarter": quarter})
    return blocks


def _build_leave_monitoring_rows(year):
    """One row per employee; per month a dict of pattern counts:
    monday / friday (approved-leave days landing on a Mon/Fri), hd (half-day
    leaves + half-day check-ins), early (Early check-outs), wfh (not tracked —
    always None), late (Present (Late) check-ins)."""
    employees = _active_employees()
    uids      = [str(e["_id"]) for e in employees]
    year_start = f"{year:04d}-01-01"
    year_end   = f"{year:04d}-12-31"

    leaves_by_uid = {}
    for lv in leaves_col.find({
        "user_id": {"$in": uids},
        "from_date": {"$lte": year_end}, "to_date": {"$gte": year_start},
        "status": {"$nin": ["Rejected", "Cancelled"]},
    }):
        if _leave_is_approved(lv):
            leaves_by_uid.setdefault(lv["user_id"], []).append(lv)

    def _month_counts(indicator=None, day_type=None, rec_type="checkin"):
        q = {"user_id": {"$in": uids}, "type": rec_type,
             "date": {"$gte": year_start, "$lte": year_end}}
        if indicator:
            q["status_indicator"] = indicator
        if day_type:
            q["day_type"] = day_type
        out = {}
        for rec in attendance_col.find(q, {"user_id": 1, "date": 1}):
            m = int(rec["date"][5:7])
            out.setdefault(rec["user_id"], {}).setdefault(m, 0)
            out[rec["user_id"]][m] += 1
        return out

    late_by_uid     = _month_counts(indicator="Present (Late)")
    early_by_uid    = _month_counts(indicator="Early", rec_type="checkout")
    hd_ci_by_uid    = _month_counts(day_type="half-day")

    rows = []
    for emp in sorted(employees, key=lambda e: (e.get("name") or "").lower()):
        uid = str(emp["_id"])
        emp_leaves = leaves_by_uid.get(uid, [])
        months = []
        for m in range(1, 13):
            days_in_month = calendar.monthrange(year, m)[1]
            mon = fri = hd = 0
            for lv in emp_leaves:
                is_half = lv.get("type") == "half"
                fd, td = lv.get("from_date", ""), lv.get("to_date", "")
                for d in range(1, days_in_month + 1):
                    ds = f"{year:04d}-{m:02d}-{d:02d}"
                    if not (fd <= ds <= td):
                        continue
                    wd = datetime(year, m, d).weekday()  # Mon=0 .. Sun=6
                    if is_half:
                        hd += 1
                    if wd == 0:
                        mon += 1
                    elif wd == 4:
                        fri += 1
            hd += hd_ci_by_uid.get(uid, {}).get(m, 0)
            months.append({
                "monday": mon, "friday": fri, "hd": hd,
                "early": early_by_uid.get(uid, {}).get(m, 0),
                "wfh": None,
                "late": late_by_uid.get(uid, {}).get(m, 0),
            })
        rows.append({
            "employee_code": emp.get("employee_code") or "—",
            "name":          emp.get("name") or "",
            "department":    _dept_of(emp),
            "months":        months,
        })
    return rows


def _build_late_coming_rows(year):
    """One row per employee: a yearly `count` and, per month, the list of
    dd/mm/yy dates the employee was stamped 'Present (Late)' at check-in."""
    employees = _active_employees()
    uids      = [str(e["_id"]) for e in employees]
    year_start = f"{year:04d}-01-01"
    year_end   = f"{year:04d}-12-31"

    by_uid = {}
    for rec in attendance_col.find({
        "user_id": {"$in": uids}, "type": "checkin",
        "status_indicator": "Present (Late)",
        "date": {"$gte": year_start, "$lte": year_end},
    }, {"user_id": 1, "date": 1}).sort("date", 1):
        m = int(rec["date"][5:7])
        by_uid.setdefault(rec["user_id"], {}).setdefault(m, []).append(rec["date"])

    rows = []
    for emp in sorted(employees, key=lambda e: (e.get("name") or "").lower()):
        uid = str(emp["_id"])
        mdict = by_uid.get(uid, {})
        months, total = [], 0
        for m in range(1, 13):
            dates = mdict.get(m, [])
            total += len(dates)
            months.append([f"{d[8:10]}/{d[5:7]}/{d[2:4]}" for d in dates])
        rows.append({
            "employee_code": emp.get("employee_code") or "—",
            "name":          emp.get("name") or "",
            "department":    _dept_of(emp),
            "count":         total,
            "months":        months,
        })
    return rows


# ── Master Tracker PDF (4 sections) ─────────────────────────────────────────

_PDF_PAGE_W_MM = max(
    30 * 3 + 15 * 12 * 4 + 22,   # Master Sheet
    38 * 3 + 24 * 12 + 20,       # Dep wise
    30 * 3 + 9 * 12 * 6 + 20,    # Leave Monitoring
    40 * 4 + 26 * 12 + 20,       # Late Coming
)

_BASE_TABLE_STYLE = [
    ("ALIGN",  (0, 0), (-1, -1), "CENTER"),
    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ("TOPPADDING", (0, 0), (-1, -1), 1.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
    ("LEFTPADDING", (0, 0), (-1, -1), 1.5), ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
]


def _master_tracker_flowables(rows, year, doc):
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    sub_labels = ["Total", "Attended", "Leaves", "%"]
    fixed_labels  = ["Employee ID", "Employee Name", "Department"]
    weights = [1.0, 1.8, 1.5] + [0.6] * (12 * 4) + [1.0]
    total_weight = sum(weights)
    col_widths = [doc.width * (w / total_weight) for w in weights]

    hdr_style  = ParagraphStyle("hdr", fontSize=6, leading=7, alignment=1, fontName="Helvetica-Bold")
    cell_style = ParagraphStyle("cell", fontSize=5.5, leading=6.5, alignment=1)

    row1 = [""] * 3 + sum(([Paragraph(calendar.month_abbr[m], hdr_style)] + [""] * 3 for m in range(1, 13)), []) + [Paragraph("Total %", hdr_style)]
    row2 = [Paragraph(l, hdr_style) for l in fixed_labels] + [Paragraph(l, hdr_style) for _ in range(12) for l in sub_labels] + [""]

    body_rows = []
    for r in rows:
        row = [Paragraph(r["employee_code"], cell_style), Paragraph(r["name"], cell_style), Paragraph(r["department"], cell_style)]
        for mo in r["months"]:
            row.append(Paragraph(f'{mo["total"]:g}' if mo["total"] else "", cell_style))
            row.append(Paragraph(f'{mo["attended"]:g}' if mo["total"] else "", cell_style))
            row.append(Paragraph(f'{mo["leaves"]:g}' if mo["total"] else "", cell_style))
            pct_style = cell_style
            if mo["pct"] is not None and mo["pct"] < 90:
                pct_style = ParagraphStyle("pctlow", parent=cell_style, textColor=rl_colors.HexColor("#dc2626"), fontName="Helvetica-Bold")
            row.append(Paragraph(f'{mo["pct"]}' if mo["pct"] is not None else "", pct_style))
        row.append(Paragraph(f'{r["year_pct"]}' if r["year_pct"] is not None else "", ParagraphStyle("yr", parent=cell_style, fontName="Helvetica-Bold")))
        body_rows.append(row)

    table = Table([row1, row2] + body_rows, colWidths=col_widths, repeatRows=2)
    style = list(_BASE_TABLE_STYLE) + [
        ("GRID",   (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 1), rl_colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0, 0), (-1, 1), rl_colors.white),
        ("ROWBACKGROUNDS", (0, 2), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
        ("SPAN", (0, 0), (0, 1)), ("SPAN", (1, 0), (1, 1)), ("SPAN", (2, 0), (2, 1)), ("SPAN", (-1, 0), (-1, 1)),
    ]
    for m in range(12):
        start = 3 + m * 4
        style.append(("SPAN", (start, 0), (start + 3, 0)))
    table.setStyle(TableStyle(style))

    title = Paragraph(f"<b>Attendance Master Tracker — {year}</b>", getSampleStyleSheet()["Heading3"])
    return [title, table]


def _dept_wise_flowables(blocks, year, doc):
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    weights = [0.7, 1.2, 2.2] + [1.0] * 12
    total_weight = sum(weights)
    col_widths = [doc.width * (w / total_weight) for w in weights]

    hdr_style  = ParagraphStyle("dhdr", fontSize=7, leading=8, alignment=1, fontName="Helvetica-Bold")
    cell_style = ParagraphStyle("dcell", fontSize=6.5, leading=7.5, alignment=1)
    sec_style  = ParagraphStyle("dsec", fontSize=7.5, leading=9, alignment=0, fontName="Helvetica-Bold", textColor=rl_colors.white)
    agg_style  = ParagraphStyle("dagg", parent=cell_style, fontName="Helvetica-Bold")

    header = [Paragraph("Sl No", hdr_style), Paragraph("Employee ID", hdr_style), Paragraph("Name", hdr_style)] + \
             [Paragraph(f"{calendar.month_abbr[m]} %", hdr_style) for m in range(1, 13)]
    data = [header]
    section_rows, agg_rows = [], []

    for block in blocks:
        section_rows.append(len(data))
        data.append([Paragraph(block["department"].upper(), sec_style)] + [""] * 14)
        for i, e in enumerate(block["employees"], 1):
            data.append([Paragraph(str(i), cell_style), Paragraph(e["employee_code"], cell_style), Paragraph(e["name"], cell_style)] +
                        [Paragraph(f'{p:g}' if p is not None else "", cell_style) for p in e["pcts"]])
        agg_rows.append(len(data))
        data.append([Paragraph("", agg_style), Paragraph("", agg_style), Paragraph("Average %", agg_style)] +
                    [Paragraph(f'{a:g}' if a is not None else "", agg_style) for a in block["average"]])
        agg_rows.append(len(data))
        data.append([Paragraph("", agg_style), Paragraph("", agg_style), Paragraph("Quarter %", agg_style)] +
                    [Paragraph(f'{q:g}' if q is not None else "", agg_style) for q in block["quarter"]])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style = list(_BASE_TABLE_STYLE) + [
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), rl_colors.white),
    ]
    for sr in section_rows:
        style.append(("BACKGROUND", (0, sr), (-1, sr), rl_colors.HexColor("#b8860b")))
        style.append(("SPAN", (0, sr), (-1, sr)))
        style.append(("ALIGN", (0, sr), (0, sr), "LEFT"))
    for ar in agg_rows:
        style.append(("BACKGROUND", (0, ar), (-1, ar), rl_colors.HexColor("#e2e8f0")))
    table.setStyle(TableStyle(style))

    title = Paragraph(f"<b>Department-wise Attendance % — {year}</b>", getSampleStyleSheet()["Heading3"])
    return [title, table]


def _leave_monitoring_flowables(rows, year, doc):
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    subs = ["Mon\nLeave", "Fri\nLeave", "HD", "Early\nLeaving", "WFH", "Late\nComing"]
    weights = [1.0, 1.8, 1.4] + [0.6] * (12 * 6)
    total_weight = sum(weights)
    col_widths = [doc.width * (w / total_weight) for w in weights]

    hdr_style  = ParagraphStyle("lhdr", fontSize=5.5, leading=6, alignment=1, fontName="Helvetica-Bold")
    cell_style = ParagraphStyle("lcell", fontSize=5.5, leading=6.5, alignment=1)

    row1 = [""] * 3 + sum(([Paragraph(calendar.month_name[m], hdr_style)] + [""] * 5 for m in range(1, 13)), [])
    row2 = [Paragraph(l, hdr_style) for l in ["Employee ID", "Name", "Department"]] + \
           [Paragraph(s.replace("\n", " "), hdr_style) for _ in range(12) for s in subs]

    body_rows = []
    for r in rows:
        row = [Paragraph(r["employee_code"], cell_style), Paragraph(r["name"], cell_style), Paragraph(r["department"], cell_style)]
        for mo in r["months"]:
            for key in ("monday", "friday", "hd", "early", "wfh", "late"):
                v = mo[key]
                row.append(Paragraph("" if not v else f"{v:g}", cell_style))
        body_rows.append(row)

    table = Table([row1, row2] + body_rows, colWidths=col_widths, repeatRows=2)
    style = list(_BASE_TABLE_STYLE) + [
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 1), rl_colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0, 0), (-1, 1), rl_colors.white),
        ("ROWBACKGROUNDS", (0, 2), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
        ("SPAN", (0, 0), (0, 1)), ("SPAN", (1, 0), (1, 1)), ("SPAN", (2, 0), (2, 1)),
    ]
    for m in range(12):
        start = 3 + m * 6
        style.append(("SPAN", (start, 0), (start + 5, 0)))
    table.setStyle(TableStyle(style))

    title = Paragraph(f"<b>Leave Pattern Monitoring — {year}</b>", getSampleStyleSheet()["Heading3"])
    return [title, table]


def _late_coming_flowables(rows, year, doc):
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    weights = [1.0, 1.8, 1.4, 0.7] + [1.6] * 12
    total_weight = sum(weights)
    col_widths = [doc.width * (w / total_weight) for w in weights]

    hdr_style  = ParagraphStyle("xhdr", fontSize=6.5, leading=7.5, alignment=1, fontName="Helvetica-Bold")
    cell_style = ParagraphStyle("xcell", fontSize=5.5, leading=7, alignment=1)

    header = [Paragraph(l, hdr_style) for l in ["Employee ID", "Name", "Department", "Count"]] + \
             [Paragraph(calendar.month_abbr[m], hdr_style) for m in range(1, 13)]
    body_rows = []
    for r in rows:
        row = [Paragraph(r["employee_code"], cell_style), Paragraph(r["name"], cell_style),
               Paragraph(r["department"], cell_style),
               Paragraph(str(r["count"]) if r["count"] else "", ParagraphStyle("xc", parent=cell_style, fontName="Helvetica-Bold"))]
        for dates in r["months"]:
            row.append(Paragraph("<br/>".join(dates), cell_style))
        body_rows.append(row)

    table = Table([header] + body_rows, colWidths=col_widths, repeatRows=1)
    style = list(_BASE_TABLE_STYLE) + [
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), rl_colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    table.setStyle(TableStyle(style))

    title = Paragraph(f"<b>Late Coming — {year}</b>", getSampleStyleSheet()["Heading3"])
    return [title, table]


def _render_master_tracker_pdf(tracker_rows, dept_blocks, leave_rows, late_rows, year):
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, PageBreak

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=(_PDF_PAGE_W_MM * mm, 297 * mm),
        leftMargin=6 * mm, rightMargin=6 * mm, topMargin=8 * mm, bottomMargin=8 * mm,
    )
    story = []
    story += _master_tracker_flowables(tracker_rows, year, doc)
    story.append(PageBreak())
    story += _dept_wise_flowables(dept_blocks, year, doc)
    story.append(PageBreak())
    story += _leave_monitoring_flowables(leave_rows, year, doc)
    story.append(PageBreak())
    story += _late_coming_flowables(late_rows, year, doc)
    doc.build(story)
    buf.seek(0)
    return buf


def _build_master_tracker_bundle(year):
    """All four sections of the Master Tracker report in one call."""
    tracker_rows = _build_master_tracker_rows(year)
    return {
        "tracker": tracker_rows,
        "dept":    _dept_wise_blocks_from_tracker(tracker_rows),
        "leave":   _build_leave_monitoring_rows(year),
        "late":    _build_late_coming_rows(year),
    }


@bp.route("/api/admin/reports/master-tracker-pdf", methods=["GET"])
@token_required
def master_tracker_pdf():
    if not _reports_allowed(request.user):
        return "Unauthorized", 403
    try:
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return "year is required", 400
    if year < 2000:
        return "Invalid year", 400

    b = _build_master_tracker_bundle(year)
    buf = _render_master_tracker_pdf(b["tracker"], b["dept"], b["leave"], b["late"], year)
    filename = f"Attendance_Master_Tracker_{year}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


def _render_master_tracker_csv(rows):
    buf = io.StringIO()
    w = csv.writer(buf)
    header = ["Employee ID", "Employee Name", "Department"]
    for m in range(1, 13):
        name = calendar.month_abbr[m]
        header += [f"{name} Total", f"{name} Attended", f"{name} Leaves", f"{name} %"]
    header.append("Total %")
    w.writerow(header)
    for r in rows:
        row = [r["employee_code"], r["name"], r["department"]]
        for mo in r["months"]:
            row += [
                f'{mo["total"]:g}' if mo["total"] else "",
                f'{mo["attended"]:g}' if mo["total"] else "",
                f'{mo["leaves"]:g}' if mo["total"] else "",
                mo["pct"] if mo["pct"] is not None else "",
            ]
        row.append(r["year_pct"] if r["year_pct"] is not None else "")
        w.writerow(row)
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    out.seek(0)
    return out


@bp.route("/api/admin/reports/master-tracker-csv", methods=["GET"])
@token_required
def master_tracker_csv():
    if not _reports_allowed(request.user):
        return "Unauthorized", 403
    try:
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return "year is required", 400
    if year < 2000:
        return "Invalid year", 400

    rows = _build_master_tracker_rows(year)
    buf  = _render_master_tracker_csv(rows)
    filename = f"Attendance_Master_Tracker_{year}.csv"
    return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


# ── Master Tracker Excel (.xlsx, one worksheet per section) ─────────────────

def _render_master_tracker_xlsx(bundle, year):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1E293B")
    SEC_FILL = PatternFill("solid", fgColor="B8860B")
    AGG_FILL = PatternFill("solid", fgColor="E2E8F0")
    CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ABBR     = [calendar.month_abbr[m] for m in range(1, 13)]
    NAMES    = [calendar.month_name[m] for m in range(1, 13)]

    def num(v):
        return "" if v in (None, 0) else v

    def style_row(ws, r, fill=HDR_FILL, font=HDR_FONT):
        for cell in ws[r]:
            cell.font = font
            cell.fill = fill
            cell.alignment = CENTER

    def autofit(ws, cap=34):
        for col in ws.columns:
            first = col[0]
            if not hasattr(first, "column"):
                continue
            longest = 0
            for c in col:
                for line in str(c.value or "").split("\n"):
                    longest = max(longest, len(line))
            ws.column_dimensions[get_column_letter(first.column)].width = min(max(longest + 2, 9), cap)

    wb = Workbook()

    # ── Sheet 1 — Master Sheet ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Master Sheet"
    header = ["Employee ID", "Employee Name", "Department"]
    for name in ABBR:
        header += [f"{name} Total", f"{name} Attended", f"{name} Leaves", f"{name} %"]
    header.append("Total %")
    ws.append(header)
    for r in bundle["tracker"]:
        row = [r["employee_code"], r["name"], r["department"]]
        for mo in r["months"]:
            row += [
                num(mo["total"]), num(mo["attended"]), num(mo["leaves"]),
                mo["pct"] if mo["pct"] is not None else "",
            ]
        row.append(r["year_pct"] if r["year_pct"] is not None else "")
        ws.append(row)
    style_row(ws, 1)
    ws.freeze_panes = "D2"
    autofit(ws)

    # ── Sheet 2 — Dep wise ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Dep wise")
    head2 = ["Sl No", "Employee ID", "Name"] + [f"{a} %" for a in ABBR]
    ws2.append(head2)
    style_row(ws2, 1)
    for block in bundle["dept"]:
        ws2.append([block["department"].upper()] + [""] * (len(head2) - 1))
        r_sec = ws2.max_row
        ws2.merge_cells(start_row=r_sec, start_column=1, end_row=r_sec, end_column=len(head2))
        style_row(ws2, r_sec, fill=SEC_FILL)
        for i, e in enumerate(block["employees"], 1):
            ws2.append([i, e["employee_code"], e["name"]] +
                       [p if p is not None else "" for p in e["pcts"]])
        ws2.append(["", "", "Average %"] + [a if a is not None else "" for a in block["average"]])
        style_row(ws2, ws2.max_row, fill=AGG_FILL, font=Font(bold=True))
        ws2.append(["", "", "Quarter %"] + [q if q is not None else "" for q in block["quarter"]])
        style_row(ws2, ws2.max_row, fill=AGG_FILL, font=Font(bold=True))
        ws2.append([])
    ws2.freeze_panes = "D2"
    autofit(ws2)

    # ── Sheet 3 — Leave Monitoring ────────────────────────────────────────
    ws3 = wb.create_sheet("Leave Monitoring")
    SUBS = ["Monday Leave", "Friday Leave", "HD", "Early Leaving", "WFH", "Late Coming"]
    row1 = ["Employee ID", "Name", "Department"]
    row2 = ["", "", ""]
    for name in NAMES:
        row1 += [name] + [""] * (len(SUBS) - 1)
        row2 += SUBS
    ws3.append(row1)
    ws3.append(row2)
    col = 4
    for _ in range(12):
        ws3.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(SUBS) - 1)
        col += len(SUBS)
    for c in ("A", "B", "C"):
        ws3.merge_cells(f"{c}1:{c}2")
    for r in bundle["leave"]:
        row = [r["employee_code"], r["name"], r["department"]]
        for mo in r["months"]:
            row += [num(mo["monday"]), num(mo["friday"]), num(mo["hd"]),
                    num(mo["early"]), "", num(mo["late"])]
        ws3.append(row)
    style_row(ws3, 1)
    style_row(ws3, 2)
    ws3.freeze_panes = "D3"
    autofit(ws3, cap=16)

    # ── Sheet 4 — Late Coming ────────────────────────────────────────────
    ws4 = wb.create_sheet("Late Coming")
    head4 = ["Employee ID", "Name", "Department", "Count"] + ABBR
    ws4.append(head4)
    style_row(ws4, 1)
    for r in bundle["late"]:
        ws4.append([r["employee_code"], r["name"], r["department"], num(r["count"])] +
                   ["\n".join(dates) for dates in r["months"]])
    for row in ws4.iter_rows(min_row=2, min_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws4.freeze_panes = "E2"
    autofit(ws4, cap=18)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route("/api/admin/reports/master-tracker-xlsx", methods=["GET"])
@token_required
def master_tracker_xlsx():
    if not _reports_allowed(request.user):
        return "Unauthorized", 403
    try:
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return "year is required", 400
    if year < 2000:
        return "Invalid year", 400

    bundle = _build_master_tracker_bundle(year)
    buf = _render_master_tracker_xlsx(bundle, year)
    filename = f"Attendance_Master_Tracker_{year}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename,
    )
